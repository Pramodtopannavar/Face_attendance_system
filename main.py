from tkinter import *
from tkinter import messagebox, filedialog,ttk
from tkinter import filedialog
import ast
import cv2
import face_recognition
import os
import numpy as np
from numpy import delete
from datetime import datetime, date
import openpyxl, xlrd
import pandas as pd
from openpyxl import Workbook,load_workbook
import pathlib
from tkinter.ttk import Combobox
from PIL import Image, ImageTk
import maskpass
from click import open_file
import tkinter as tk

root = Tk()
root.title('login')
root.geometry('925x500+300+200')
root.configure(bg="#fff")
root.resizable(False, False)


def signup():
    window = Toplevel()
    window.title("signup")
    window.geometry('925x500+300+200')
    window.configure(bg="#fff")
    window.resizable(False, False)

    def signup1():
        username = user1.get()
        password = code1.get()
        confirm_password = code2.get()
        if password == confirm_password:
            try:
                file = open('sigup.txt', 'r+')
                d = file.read()
                r1 = ast.literal_eval(d)
                dict = {username: password}
                r1.update(dict)
                file.truncate(0)
                file.close()
                file = open('sigup.txt', 'w')
                w = file.write(str(r1))
                messagebox.showinfo('signup1', 'successfully sign up')
            except:
                file = open('sigup.txt', 'w')
                pp = str({'username': 'password'})
                file.write(pp)
                file.close()
        else:
            messagebox.showerror('Invalid', "Account not created")

    def sign():
        window.destroy()

    box = Frame(window, width=350, height=390, bg='#fff')
    box.place(x=480, y=50)
    img = PhotoImage(file='12.png')
    Label(window, image=img, bg='white').place(x=50, y=60)
    heading1 = Label(box, text='sign up', fg="#57a1f8", bg='white', font=('Microsoft YaHei UI Light', 23, 'bold'))
    heading1.place(x=100, y=5)

    def on_enter1(e):
        user1.delete(0, 'end')

    def on_leave1(e):

        if user1.get() == '':
            user1.insert(0, 'Username')

    user1 = Entry(box, width=25, fg='black', border=2, bg='white', font=('Microsoft YaHei UI Light', 11))
    user1.place(x=30, y=80)
    user1.insert(0, 'Username')
    user1.bind('<FocusIn>', on_enter1)
    user1.bind('<FocusOut>', on_leave1)

    Frame(box, width=295, height=2, bg='black').place(x=25, y=107)

    def on_enter1(e):
        code1.delete(0, 'end')

    def on_leave1(e):

        if code1.get() == '':
            code1.insert(0, 'password')

    code1 = Entry(box, width=25, fg='black', border=2, bg='white', font=('Microsoft YaHei UI Light', 11))
    code1.place(x=30, y=150)
    code1.insert(0, 'password')
    code1.bind('<FocusIn>', on_enter1)
    code1.bind('<FocusOut>', on_leave1)

    Frame(box, width=295, height=2, bg='black').place(x=25, y=177)

    def on_enter1(e):
        code2.delete(0, 'end')

    def on_leave1(e):

        if code2.get() == '':
            code2.insert(0, 'confirm password')

    code2 = Entry(box, width=25, fg='black', border=2, bg='white', font=('Microsoft YaHei UI Light', 11))
    code2.place(x=30, y=220)
    code2.insert(0, ' confirm password')
    code2.bind('<FocusIn>', on_enter1)
    code2.bind('<FocusOut>', on_leave1)

    Frame(box, width=295, height=2, bg='black').place(x=25, y=247)

    Button(box, width=39, pady=7, text='sign up', bg='#57a1f8', fg='white', border=0, command=signup1).place(x=35,
                                                                                                             y=280)
    label1 = Label(box, text='I have a account', fg='black', bg='white', font=('Microsoft YaHei UI Light', 9))
    label1.place(x=90, y=340)

    Button(box, width=6, text='sign in', border=0, bg='white', cursor='hand2', fg='#57a1f8', command=sign).place(x=200,
                                                                                                                 y=340)

    window.mainloop()


def signin():
    username = user.get()
    password = code.get()

    file = open('sigup.txt', 'r')
    d = file.read()
    r = ast.literal_eval(d)
    file.close()

    print(r.keys())
    print(r.values())
    if username in r.keys() and password == r[username]:
        root.destroy()
        background = "black"
        framebg = "#EDEDED"
        framefg = "#06283D"

        screen = Tk()
        screen.title("App")
        screen.geometry('925x500+300+200')
        screen.config(bg=background)





        # Exit
        def Exit():
            screen.destroy()

        ###update
        def Upadate():

            background = "#06283D"
            framebg = "#EDEDED"
            framefg = "#06283D"
            upd = Toplevel()
            upd.title("student Registration system")
            upd.geometry("1250x700+210+100")
            upd.config(bg=background)

            def serch():

                txt = Search.get()

                file = openpyxl.load_workbook("studentdata.xlsx")
                sheet = file.active
                for row in sheet.rows:
                    if row[0].value == int(txt):
                        name = row[0]
                        ###print(str(name))
                        reg_no_position = str(name)[14:-1]
                        reg_number = str(name)[15:-1]

                    ### print(reg_no_position)
                    ####print(reg_number)
                try:
                    print(str(name))

                except:
                    messagebox.showerror("Invalid", "Invalid  registrstion number!!!")
                x1 = sheet.cell(row=int(reg_number), column=1).value
                x2 = sheet.cell(row=int(reg_number), column=2).value
                x3 = sheet.cell(row=int(reg_number), column=3).value
                x4 = sheet.cell(row=int(reg_number), column=4).value
                x5 = sheet.cell(row=int(reg_number), column=5).value
                x6 = sheet.cell(row=int(reg_number), column=6).value
                x7 = sheet.cell(row=int(reg_number), column=7).value

                # print(x1)

                Registration.set(x1)
                Name.set(x2)
                sem.set(x3)

                if x4 == 'female':
                    R2.select()
                else:
                    R1.select()

                DOB.set(x5)
                Usn.set(x6)
                Date.set(x7)
                img10 = (Image.open(
                    "C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images\\" + str(x2) + ".jpg"))
                resized_image = img10.resize((190, 190))
                photo2 = ImageTk.PhotoImage(resized_image)
                lb1.config(image=photo2)
                lb1.image = photo2

            def showimg5():
                global img11

                filename5 = filedialog.askopenfilename(initialdir=os.getcwd(),
                                                       title="select image file")
                img11 = (Image.open(filename5))
                resized_image = img11.resize((190, 190))
                photo2 = ImageTk.PhotoImage(resized_image)
                lb1.config(image=photo2)
                lb1.image = photo2

            def Exit():
                upd.destroy()

            def clear():
                global img8
                Name.set('')
                DOB.set('')
                sem.set("select semester")
                Usn.set('')

                saveButton.config(state='normal')
                img1 = PhotoImage(file='image/uploadphoto.png')
                lb1.config(image=img1)
                lb1.image = img1
                img11 = ""

            def update1():
                R1 = Registration.get()
                N1 = Name.get()
                U1 = Usn.get()
                selection()
                G1 = gender
                D2 = DOB.get()
                D1 = Date.get()
                S1 = sem.get()

                file = openpyxl.load_workbook("studentdata.xlsx")
                sheet = file.active
                for row in sheet.rows:
                    if row[0].value == R1:
                        name = row[0]
                        print(str(name))
                        reg_no_position = str(name)[14:1]
                        reg_number = str(name)[15:-1]
                        print(reg_number)
                # sheet.cell(column=1,row=int(reg_number),value=R1)
                sheet.cell(column=2, row=int(reg_number), value=N1)
                sheet.cell(column=3, row=int(reg_number), value=S1)
                sheet.cell(column=4, row=int(reg_number), value=G1)
                sheet.cell(column=5, row=int(reg_number), value=D2)
                sheet.cell(column=6, row=int(reg_number), value=U1)
                sheet.cell(column=7, row=int(reg_number), value=D1)
                file.save(r'studentdata.xlsx')
                try:
                    img11.save("student_images/" + str(N1) + ".jpg")
                except:
                    pass
                messagebox.showinfo("update", "update Succefully!!")
                clear()

            def selection():
                global gender
                value = radio.get()
                if value == 1:
                    gender = "male"

                else:
                    gender = "Female"

            Label(upd, text="Email: pramodtopannavar843@gmail.com", width=10, height=3, bg="yellow", anchor='e').pack(
                side=TOP, fill=X)
            Label(upd, text="Update", width=10, height=2, bg="light green", fg='black',
                  font='arial 20 bold').pack(side=TOP, fill=X)
            Search = StringVar()
            Entry(upd, textvariable=Search, width=15, bd=2, font="arial 20").place(x=820, y=70)
            imageicon3 = PhotoImage(file="image/search.png")
            Srch = Button(upd, text="search", compound=LEFT, image=imageicon3, width=121, bg='#68ddfa',
                          font="arial 13 bold", command=serch)
            Srch.place(x=1060, y=64)
            Label(upd, text="Registration no:", font="arial 13", fg=framebg, bg=background).place(x=30, y=150)
            Label(upd, text="Date:", font="arial 13", fg=framebg, bg=background).place(x=500, y=150)

            Registration = IntVar()
            Date = StringVar()

            reg_entry = Entry(upd, textvariable=Registration, width=15, font="arial 10").place(x=160, y=150)

            today = date.today()
            d1 = today.strftime("%d/%m/%y")
            date_entry = Entry(upd, textvariable=Date, width=15, font="arial 10")
            date_entry.place(x=550, y=150)

            Date.set(d1)

            # student details
            p = LabelFrame(upd, text="student's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=400,
                           relief=GROOVE)
            p.place(x=30, y=200)

            Label(p, text="Full name:", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=50)
            Label(p, text="USN:", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=100)
            Label(p, text="gender:", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=150)
            Label(p, text="Semester:", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=200)
            Label(p, text="Date of birth", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=250)

            Name = StringVar()
            name_entry = Entry(p, textvariable=Name, width=20, font="arial 10")
            name_entry.place(x=160, y=50)

            Usn = StringVar()
            usn_entry = Entry(p, textvariable=Usn, width=20, font="arial 10")
            usn_entry.place(x=160, y=100)

            radio = IntVar()
            R1 = Radiobutton(p, text="MALE", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
            R1.place(x=150, y=150)
            R2 = Radiobutton(p, text="FEMALE", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
            R2.place(x=250, y=150)

            sem = Combobox(p, values=['3', '4', '5', '6', '7', '8'], font="Roboto 10", width=17, state="r")
            sem.place(x=160, y=200)
            sem.set("select semester")

            DOB = StringVar()
            dob_entry = Entry(p, textvariable=DOB, width=20, font="arial 10")
            dob_entry.place(x=160, y=250)

            # image
            f = Frame(upd, bd=3, bg="black", width=200, height=200, relief=GROOVE)
            f.place(x=1000, y=150)

            imgu = PhotoImage(file="image/uploadphoto.png")
            lb1 = Label(upd, bg="black", image=imgu)
            lb1.place(x=1000, y=150)

            # button
            Button(upd, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue",
                   command=showimg5).place(x=1000, y=370)
            saveButton = Button(upd, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen",
                                command=update1)
            saveButton.place(x=1000, y=450)
            Button(upd, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=clear).place(
                x=1000, y=530)
            Button(upd, text="Exit", width=19, height=2, font="arial 12 bold", bg="grey", command=Exit).place(x=1000,
                                                                                                              y=610)

            mainloop()

        def enroll():


            background = "#06283D"
            framebg = "#EDEDED"
            framefg = "#06283D"
            reg = Toplevel()
            reg.title("student Registration system")
            reg.geometry("1250x700+210+100")
            reg.config(bg=background)
            file1 = pathlib.Path('student_data.xlsx')


            ################## Exit##################
            def Exit():
                reg.destroy()

            ##########show image######################
            def showimg5():
                global img8

                filename5 = filedialog.askopenfilename(initialdir=os.getcwd(),
                                                       title="select image file")
                img8 = (Image.open(filename5))
                resized_image = img8.resize((190, 190))
                photo2 = ImageTk.PhotoImage(resized_image)
                lb1.config(image=photo2)
                lb1.image = photo2

            ################Registration######################
            def registration_no():
                file8 = openpyxl.load_workbook('studentdata.xlsx')
                sheet = file8.active
                row = sheet.max_row
                max_row_value = sheet.cell(row=row, column=1).value

                try:
                    Registration.set(max_row_value + 1)
                except:
                    Registration.set("1")

            #####################clear#########################
            def clear():
                global img8
                Name.set('')
                DOB.set('')
                sem.set("select semester")
                Usn.set('')
                registration_no()
                saveButton.config(state='normal')
                img1 = PhotoImage(file='image/uploadphoto.png')
                lb1.config(image=img1)
                lb1.image = img1
                img8 = ""

            ##################save#######################
            def save():
                R1 = Registration.get()
                N1 = Name.get()
                U1 = Usn.get()
                try:
                    G1 = gender
                except:
                    messagebox.showerror("error", "select Gender")
                D2 = DOB.get()
                D1 = Date.get()
                S1 = sem.get()

                if N1 == "" or S1 == "select semester" or U1 == "" or D2 == "":
                    messagebox.showerror("error", "Few Data is missing")
                else:
                    file = openpyxl.load_workbook('studentdata.xlsx')
                    sheet = file.active
                    sheet.cell(column=1, row=sheet.max_row + 1, value=R1)
                    sheet.cell(column=2, row=sheet.max_row, value=N1)
                    sheet.cell(column=3, row=sheet.max_row, value=S1)
                    sheet.cell(column=4, row=sheet.max_row, value=G1)
                    sheet.cell(column=5, row=sheet.max_row, value=D2)
                    sheet.cell(column=6, row=sheet.max_row, value=U1)
                    sheet.cell(column=7, row=sheet.max_row, value=D1)
                    file.save(r'studentdata.xlsx')
                    try:
                        img8.save("student_images/" + str(N1) + ".jpg")
                    except:
                        messagebox.showerror("info", "profile picture is not available!!!!")
                    messagebox.showinfo("info", "sucessfully data entered!!!!")
                    clear()
                    registration_no()




                ##########gender

            def selection():
                global gender
                value = radio.get()
                if value == 1:
                    gender = "male"

                else:
                    gender = "Female"

            # TOP FRAME
            Label(reg, text="Email: pramodtopannavar843@gmail.com", width=10, height=3, bg="#f0687c", anchor='e').pack(
                side=TOP, fill=X)
            Label(reg, text="STUDENT REGISTRATION", width=10, height=2, bg="#c36464", fg='#fff',
                  font='arial 20 bold').pack(side=TOP, fill=X)



            # registrstion and dataa
            Label(reg, text="Registration no:", font="arial 13", fg=framebg, bg=background).place(x=30, y=150)
            Label(reg, text="Date:", font="arial 13", fg=framebg, bg=background).place(x=500, y=150)

            Registration = IntVar()
            Date = StringVar()

            reg_entry = Entry(reg, textvariable=Registration, width=15, font="arial 10").place(x=160, y=150)

            registration_no()

            today = date.today()
            d1 = today.strftime("%d/%m/%y")
            date_entry = Entry(reg, textvariable=Date, width=15, font="arial 10")
            date_entry.place(x=550, y=150)

            Date.set(d1)

            # student details
            p = LabelFrame(reg, text="student's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=400,
                           relief=GROOVE)
            p.place(x=30, y=200)

            Label(p, text="Full name:", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=50)
            Label(p, text="USN:", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=100)
            Label(p, text="gender:", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=150)
            Label(p, text="Semester:", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=200)
            Label(p, text="Date of birth", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=250)

            Name = StringVar()
            name_entry = Entry(p, textvariable=Name, width=20, font="arial 10")
            name_entry.place(x=160, y=50)

            Usn = StringVar()
            usn_entry = Entry(p, textvariable=Usn, width=20, font="arial 10")
            usn_entry.place(x=160, y=100)

            radio = IntVar()
            R1 = Radiobutton(p, text="MALE", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
            R1.place(x=150, y=150)
            R2 = Radiobutton(p, text="FEMALE", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
            R2.place(x=250, y=150)

            sem = Combobox(p, values=['3', '4', '5', '6', '7', '8'], font="Roboto 10", width=17, state="r")
            sem.place(x=160, y=200)
            sem.set("select semester")

            DOB = StringVar()
            dob_entry = Entry(p, textvariable=DOB, width=20, font="arial 10")
            dob_entry.place(x=160, y=250)

            # image
            f = Frame(reg, bd=3, bg="black", width=200, height=200, relief=GROOVE)
            f.place(x=1000, y=150)

            imgu = PhotoImage(file="image/uploadphoto.png")
            lb1 = Label(reg, bg="black", image=imgu)
            lb1.place(x=1000, y=150)

            # button
            Button(reg, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue",
                   command=showimg5).place(x=1000, y=370)
            saveButton = Button(reg, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen",
                                command=save)
            saveButton.place(x=1000, y=450)
            Button(reg, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=clear).place(
                x=1000, y=530)
            Button(reg, text="Exit", width=19, height=2, font="arial 12 bold", bg="grey", command=Exit).place(x=1000,
                                                                                                              y=610)

            mainloop()

        def information():
            background="sky blue"
            infor=Tk()
            infor.title("Excel datasheet viewer")
            infor.geometry('1100x400+200+200')
            infor.config(bg=background)
            def D():
                infor.destroy()

            tree=ttk.Treeview(infor)
            tree.pack()




            filename='C:\\Users\Asus\\PycharmProjects\\face attendance app\\studentdata.xlsx'
            filename=r"{}".format(filename)
            df=pd.read_excel(filename)
            tree['column']=list(df.columns)
            tree['show']="headings"

            for col in tree['column']:
                tree.heading(col,text=col)

            df_rows=df.to_numpy().tolist()
            for row in df_rows:
                tree.insert("","end",values=row)


            Button(infor, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=D
                   ).place(x=800, y=250)

        def DET():
            messagebox.showinfo("DEVELOPED BY","PRAMOD T\nPRADEEP M\nPRABHUGOUDA\nRAJU B\nSHASHIKUMAR N\nATMANAND\nREVATI K\nSAKSHI")
        def abot():
            bt=Tk()
            bt.title("About")
            bt.geometry('500x480+200+200')
            background = "black"
            bt.config(bg=background)
            Label(bt, text="Welcome to our Face Attendance! Our System is designed to\n make attendance tracking easier and accurate using\n facial recognition technology.", font="arial 13", fg=framefg, bg=framebg).place(x=5, y=50)
            Label(bt, text="Features:       ", font="arial 13", fg=framefg, bg=framebg).place(x=5, y=150)
            Label(bt, text="->Fast and accurate recognition of faces        \n->Easy registration of student                        \n->Easy to get  student information                 \n->Easy to take attendance and attendance list", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=180)
            Label(bt, text="How to use:       ", font="arial 13", fg=framefg, bg=framebg).place(x=5, y=280)
            Label(bt, text="1.Enroll the student information with clear photo.                ", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=310)
            Label(bt, text="2.If any correction in the enroll of student update it.            ", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=330)
            Label(bt, text="3.we can view student information list.                               ", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=350)
            Label(bt, text="4.Start taking attendance on clicking on attendance button ", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=370)
            Label(bt, text="5.You can view the attendance records on clicking the       ", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=390)
            Label(bt, text="attendance View button.                                                   ", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=410)
        def attend():
            ate=Tk()
            ate.geometry("1250x600+210+100")
            ate.title("Attendance")
            background="black"
            ate.config(bg=background)

            def takeattendance1():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance1.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance2():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance2.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance3():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance3.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance4():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance4.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance5():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance5.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance6():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance6.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance7():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance7.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance8():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance8.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance9():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance9.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance10():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance10.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance11():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance11.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance12():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance12.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance13():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance13.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance14():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance14.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance15():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance15.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance16():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance16.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance17():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance17.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance18():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance18.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance19():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance19.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance20():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance20.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance21():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance21.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance22():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance22.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance23():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance23.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance24():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance24.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance25():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance25.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance26():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance26.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance27():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance27.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance28():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance28.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance29():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance29.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance30():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance30.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def takeattendance31():
                imgelon_bgr = face_recognition.load_image_file('venv/elon.jpg')
                imgelon_rgb = cv2.cvtColor(imgelon_bgr, cv2.COLOR_BGR2RGB)

                imgelon = face_recognition.load_image_file('venv/elon.jpg')
                imgelon = cv2.cvtColor(imgelon, cv2.COLOR_BGR2RGB)
                # ----------Finding face Location for drawing bounding boxes-------
                face = face_recognition.face_locations(imgelon_rgb)[0]
                copy = imgelon.copy()

                # -------------------Drawing the Rectangle-------------------------
                cv2.rectangle(copy, (face[3], face[0]), (face[1], face[2]), (255, 0, 255), 2)

                face = face_recognition.face_locations(imgelon)[0]
                train_elon_encod = face_recognition.face_encodings(imgelon)[0]
                test = face_recognition.load_image_file('venv/elon2.jpg')
                test = cv2.cvtColor(test, cv2.COLOR_BGR2RGB)
                test_encode = face_recognition.face_encodings(test)[0]

                path = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\student_images'
                images = []
                classNames = []
                mylist = os.listdir(path)
                for cl in mylist:
                    curImg = cv2.imread(f'{path}/{cl}')
                    images.append(curImg)
                    classNames.append(os.path.splitext(cl)[0])

                def findEncodings(images):
                    encodeList = []
                    for img in images:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        encoded_face = face_recognition.face_encodings(img)[0]
                        encodeList.append(encoded_face)
                    return encodeList

                encoded_face_train = findEncodings(images)

                def markAttendance(name):
                    with open('C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance31.csv', 'r+') as f:
                        myDataList = f.readlines()
                        nameList = []
                        for line in myDataList:
                            entry = line.split(',')
                            nameList.append(entry[0])
                        if name not in nameList:
                            now = datetime.now()
                            time = now.strftime('%I:%M:%S:%p')
                            date = now.strftime('%d-%B-%Y')
                            f.writelines(f'{name}, {time}, {date}\n')

                cap = cv2.VideoCapture(0)

                while True:
                    success, img = cap.read()
                    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
                    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)
                    faces_in_frame = face_recognition.face_locations(imgS)
                    encoded_faces = face_recognition.face_encodings(imgS, faces_in_frame)
                    for encode_face, faceloc in zip(encoded_faces, faces_in_frame):
                        matches = face_recognition.compare_faces(encoded_face_train, encode_face)
                        faceDist = face_recognition.face_distance(encoded_face_train, encode_face)
                        matchIndex = np.argmin(faceDist)
                        print(matchIndex)
                        if matches[matchIndex]:
                            name = classNames[matchIndex].upper().lower()
                            y1, x2, y2, x1 = faceloc
                            # since we scaled down by 4 times
                            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                            cv2.putText(img, name, (x1 + 6, y2 - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                            markAttendance(name)
                    cv2.imshow('webcam', img)
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

            def bck():
                ate.destroy()
            Label(ate, text="Take attendance", width=10, height=2, bg="light blue", fg='black',
                  font='arial 20 bold').pack(side=TOP, fill=X)
            Button(ate, width=30, text='class 1', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance1).place(x=25, y=150)
            Button(ate, width=30, text='class 2', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance2).place(x=325, y=150)
            Button(ate, width=30, text='class 3', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance3).place(x=625, y=150)
            Button(ate, width=30, text='class 4', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance4).place(x=925, y=150)
            Button(ate, width=30, text='class 5', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance5).place(x=25, y=200)
            Button(ate, width=30, text='class 6', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance6).place(x=325, y=200)
            Button(ate, width=30, text='class 7', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance7).place(x=625, y=200)
            Button(ate, width=30, text='class 8', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance8).place(x=925, y=200)
            Button(ate, width=30, text='class 9', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance9).place(x=25, y=250)
            Button(ate, width=30, text='class 10', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance10).place(x=325, y=250)
            Button(ate, width=30, text='class 11', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance11).place(x=625, y=250)
            Button(ate, width=30, text='class 12', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance12).place(x=925, y=250)
            Button(ate, width=30, text='class 13', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance13).place(x=25, y=300)
            Button(ate, width=30, text='class 14', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance14).place(x=325, y=300)
            Button(ate, width=30, text='class 15', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance15).place(x=625, y=300)
            Button(ate, width=30, text='class 16', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance16).place(x=925, y=300)
            Button(ate, width=30, text='class 17', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance17).place(x=25, y=350)
            Button(ate, width=30, text='class 18', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance18).place(x=325, y=350)
            Button(ate, width=30, text='class 19', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance19).place(x=625, y=350)
            Button(ate, width=30, text='class 20', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance20).place(x=925, y=350)
            Button(ate, width=30, text='class 21', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance21).place(x=25, y=400)
            Button(ate, width=30, text='class 22', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance22).place(x=325, y=400)
            Button(ate, width=30, text='class 23', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance23).place(x=625, y=400)
            Button(ate, width=30, text='class 24', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance24).place(x=925, y=400)
            Button(ate, width=30, text='class 25', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance25).place(x=25, y=450)
            Button(ate, width=30, text='class 26', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance26).place(x=325, y=450)
            Button(ate, width=30, text='class 27', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance27).place(x=625, y=450)
            Button(ate, width=30, text='class 28', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance28).place(x=925, y=450)
            Button(ate, width=30, text='class 29', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance29).place(x=25, y=500)
            Button(ate, width=30, text='class 30', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance30).place(x=325, y=500)
            bt31=Button(ate, width=30, text='class 31', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=takeattendance31).place(x=625, y=500)
            Button(ate, width=30, text='Back', border=1, bg='red', cursor='hand2', fg='black',
                   command=bck).place(x=925, y=500)

        def view1():
            atev=Tk()
            atev.geometry("1250x600+210+100")
            atev.title("Attendance")
            background="black"
            atev.config(bg=background)

            def day1():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance1.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day2():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance2.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)
            def day3():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance3.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day4():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance4.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day5():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance5.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day6():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance6.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)
            def day7():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance7.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)
            def day8():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance8.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day9():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance9.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day10():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance10.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day11():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance11.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day12():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance12.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day13():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance13.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day14():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance14.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day15():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance15.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day16():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance16.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day17():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance17.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day18():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance18.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day19():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance19.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day20():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance20.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day21():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance21.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day22():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance22.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day23():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance23.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day24():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance24.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day25():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance25.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day26():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance26.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day27():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance27.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day28():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance28.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day29():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance29.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day30():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance30.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def day31():

                at = Tk()
                at.title("Excel datasheet viewer")
                at.geometry('1100x400+200+200')

                def B():
                    at.destroy()

                def V(self):
                    selected_rows = self.table.selection()
                    for row in selected_rows:
                        self.df.drop(int(row), inplace=True)

                    self.df.to_excel(self.file_path, index=False)
                    self.table.delete(*self.table.get_children())
                    for row in self.df.to_numpy().tolist():
                        self.table.insert("", "end", values=row)

                tree = ttk.Treeview(at)
                tree.pack()

                filename = 'C:\\Users\\Asus\\PycharmProjects\\face attendance app\\venv\\Attendance31.csv'
                excel_file = r"{}".format(filename)
                df1 = pd.read_csv(excel_file)
                tree['column'] = list(df1.columns)
                tree['show'] = "headings"

                for col in tree['column']:
                    tree.heading(col, text=col)

                df_rows = df1.to_numpy().tolist()
                for row in df_rows:
                    tree.insert("", "end", values=row)

                Button(at, width=20, text='back', border=1, bg='yellow', cursor='hand2', fg='black', command=B
                       ).place(x=800, y=250)

            def bck():
                atev.destroy()

            Label(atev, text="Attendance report", width=10, height=2, bg="light blue", fg='black',
                  font='arial 20 bold').pack(side=TOP, fill=X)
            Button(atev, width=30, text='class 1', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day1).place(x=25, y=150)
            Button(atev, width=30, text='class 2', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day2).place(x=325, y=150)
            Button(atev, width=30, text='class 3', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day3).place(x=625, y=150)
            Button(atev, width=30, text='class 4', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day4).place(x=925, y=150)
            Button(atev, width=30, text='class 5', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day5).place(x=25, y=200)
            Button(atev, width=30, text='class 6', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day6).place(x=325, y=200)
            Button(atev, width=30, text='class 7', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day7).place(x=625, y=200)
            Button(atev, width=30, text='class 8', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day8).place(x=925, y=200)
            Button(atev, width=30, text='class 9', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day9).place(x=25, y=250)
            Button(atev, width=30, text='class 10', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day10).place(x=325, y=250)
            Button(atev, width=30, text='class 11', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day11).place(x=625, y=250)
            Button(atev, width=30, text='class 12', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day12).place(x=925, y=250)
            Button(atev, width=30, text='class 13', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day13).place(x=25, y=300)
            Button(atev, width=30, text='class 14', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day14).place(x=325, y=300)
            Button(atev, width=30, text='class 15', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day15).place(x=625, y=300)
            Button(atev, width=30, text='class 16', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day16).place(x=925, y=300)
            Button(atev, width=30, text='class 17', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day17).place(x=25, y=350)
            Button(atev, width=30, text='class 18', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day18).place(x=325, y=350)
            Button(atev, width=30, text='class 19', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day19).place(x=625, y=350)
            Button(atev, width=30, text='class 20', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day20).place(x=925, y=350)
            Button(atev, width=30, text='class 21', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day21).place(x=25, y=400)
            Button(atev, width=30, text='class 22', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day22).place(x=325, y=400)
            Button(atev, width=30, text='class 23', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day23).place(x=625, y=400)
            Button(atev, width=30, text='class 24', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day24).place(x=925, y=400)
            Button(atev, width=30, text='class 25', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day25).place(x=25, y=450)
            Button(atev, width=30, text='class 26', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day26).place(x=325, y=450)
            Button(atev, width=30, text='class 27', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day27).place(x=625, y=450)
            Button(atev, width=30, text='class 28', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day28).place(x=925, y=450)
            Button(atev, width=30, text='class 29', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day29).place(x=25, y=500)
            Button(atev, width=30, text='class 30', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day30).place(x=325, y=500)
            Button(atev, width=30, text='class 31', border=1, bg='yellow', cursor='hand2', fg='black',
                   command=day31).place(x=625, y=500)
            Button(atev, width=30, text='Back', border=1, bg='red', cursor='hand2', fg='black',
                   command=bck).place(x=925, y=500)



        head = Label(screen, text='Welcome ', fg='red', bg="black", font=('arial 20 bold', 23, 'underline'))
        head.place(x=400, y=50)

        enroll = Button(screen, width=20, text='Enroll', border=1, bg='yellow', cursor='hand2', fg='black',
                        command=enroll).place(x=400, y=150)
        update = Button(screen, width=20, text='Update', border=1, bg='yellow', cursor='hand2', fg='black',
                        command=Upadate).place(
            x=400, y=200)
        info = Button(screen, width=20, text='Student information', border=1, bg='yellow', cursor='hand2', fg='black',command=information).place(
            x=400, y=250)
        attendance = Button(screen, width=20, text='Attendance', border=1, bg='yellow', cursor='hand2',
                            fg='black', command=attend).place(x=400, y=300)
        attendanceview = Button(screen, width=20, text='Attendance_view', border=1, bg='yellow', cursor='hand2',
                                fg='black', command=view1).place(x=400, y=350)
        exit = Button(screen, width=20, text='Exit', border=1, bg='lightgreen', cursor='hand2', fg='black',
                      command=Exit).place(x=400,
                                          y=450)
        det = Button(screen, width=10, text='DEVELOPERS', border=1, bg='white', cursor='hand2', fg='black'
                        ,command=DET).place(x=800, y=450)
        abt = Button(screen, width=10, text='About', border=1, bg='white', cursor='hand2', fg='black'
                     , command=abot).place(x=800, y=425)







    elif username in r.keys() and password != r[username]:
        messagebox.showerror("Invalid", "invalid password")

    elif username not in r.keys():
        messagebox.showerror("Invalid", "invalid username")


img = PhotoImage(file='login.png')
Label(root, image=img, bg='white').place(x=50, y=50)
frame = Frame(root, width=350, height=350, bg="white")
frame.place(x=480, y=70)
heading = Label(frame, text='sign in', fg='#57a1f8', bg="white", font=('Microsoft yaHei UI Light', 23, 'bold'))
heading.place(x=100, y=5)


def on_enter(e):
    user.delete(0, 'end')


def on_leave(e):
    name = user.get()
    if name == '':
        user.insert(0, 'Username')


user = Entry(frame, width=25, fg='black', border=2, bg="white", font=('Microsoft YaHei UI Light', 11))
user.place(x=30, y=80)
user.insert(0, 'Username')
user.bind('<FocusIn>', on_enter)
user.bind('<FocusOut>', on_leave)

Frame(frame, width=295, height=2, bg='black').place(x=25, y=107)


def on_enter(e):
    code.delete(0, 'end')


def on_leave(e):
    name = code.get()
    if name == '':
        code.insert(0, 'password')


code = Entry(frame, width=25, fg='black', border=2, bg="white",show='*', font=('Microsoft YaHei UI Light', 11))
code.place(x=30, y=150)
code.insert(0, 'Password')
code.bind('<FocusIn>', on_enter)
code.bind('<FocusOut>', on_leave)




Frame(frame, width=295, height=2, bg='black').place(x=25, y=177)

Button(frame, width=39, text='sign in', bg='#57a1f8', fg='white', border=0, command=signin).place(x=35, y=204)
label = Label(frame, text="Don't have an account?", fg='black', bg='white', font=('Microsoft YaHei UI Light', 15))
label.place(x=75, y=270)

Button(frame, width=6, text='sign up', border=0, bg='white', cursor='hand2', fg='#57a1f8',font=('Microsoft YaHei UI Light', 15), command=signup).place(x=120,
                                                                                                                 y=305)

root.mainloop()
